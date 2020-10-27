#!/usr/bin/env python
# -*- coding: cp1252 -*-


from selenium.webdriver.common.keys import Keys
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import unittest, time, re, uuid
from selenium.webdriver.common.action_chains import ActionChains
from distutils.version import StrictVersion
from numbers import Number
#import configparser as cfgparser
import socket
import selenium.webdriver.support.ui as ui
from selenium.webdriver.support import expected_conditions as EC
import re
from datetime import datetime

#import MySQLdb
# import psycopg2,psycopg2.extras

import time
import sys
import os
import random
import glob
import re

import traceback
import glob
import urllib
from datetime import date
from openpyxl.workbook import Workbook
# import importlib

reload(sys)

sys.setdefaultencoding("cp1252")
# importlib.reload(sys)





class Traitement():
    def __init__(self):
        # app=wx.App(0)
        #13 03 2018 python2.7

        try:
            lock=open("scraping_courses.lock", "a")
            lock.close()

            date_jour1 = str(date.today())
            date_jour=self.date2fr(date_jour1,"/")
            print("date : ", date_jour)
            date_traitement=self.date2fr(date_jour1,"-")
            # d={}
            # chaine=self.NettoyagePonctPresent("nétoo/\yà g'e").replace(" ","_")
            # d.setdefault("un", []).append("a")
            # d.setdefault("un", []).append("b")
            # d.setdefault("deux", []).append("c")
            # d.setdefault("deux", []).append("d")
            # for di in d:
            #     print(d[di])

        except Exception as inst:
            msgs = 'Erreur '+'\n'
            msgs +=  'type ERREUR:'+str(type(inst))+'\n'
            msgs+=  'CONTENU:'+str(inst)+'\n'
            print(msgs)
            sys.exit(0)

        try:
            k = 0

            chromeOptions = webdriver.ChromeOptions()
            chromeOptions.add_argument("--start-maximized")

            prefs = {"profile.default_content_settings.popups": 0,
                     "download.default_directory": "", # IMPORTANT - ENDING SLASH V IMPORTANT
                     "directory_upgrade": True, "extensions_to_open": "", "plugins.plugins_disabled": ["Chrome PDF Viewer"], "plugins.plugins_list": [{"enabled":False,"name":"Chrome PDF Viewer"}]}

            chromeOptions.add_experimental_option("prefs",prefs)
            chromeOptions.add_argument("--disable-print-preview")
            chromedriver = r"chromedriver.exe"
            driver = webdriver.Chrome(executable_path=chromedriver, chrome_options=chromeOptions)

            driver.implicitly_wait(60)

            wait = ui.WebDriverWait(driver,60)

            # #Precision du chemin du driver
            # phdriver = r"phantomjs.exe"
            # driver = webdriver.PhantomJS(executable_path=phdriver)
            # #Attendre 60 secondes pour chargement page
            # driver.implicitly_wait(60)
            # wait = ui.WebDriverWait(driver,60)

            #Recuperation

            # a_lien=driver.find_element_by_xpath("//div[@class='panel-body']/ul[@class='list-group']/li[1]/div/div[2]/a")
            # a_lien.click()
            # #--------R1
            rep="resultats"
            if(os.access(rep,os.F_OK)==False):
                os.makedirs(rep,777)

            if(os.access(r"code_pays.txt",os.F_OK)==False):
                print("La liste du code pays est introuvable")
                sys.exit(0)

            code_pays=[]
            with open(r"code_pays.txt", "r") as f :
                fichier_entier = f.read()
                lignes = fichier_entier.split("\n")
                code_pays=lignes

            liste1={}

            prix=""

            num_fichier=0
            l="file:///C:/Users/SI/Desktop/ParisTurf/ARRIVEES/2012/08/02/R2/01.htm"
            driver.get(l)
            driver.maximize_window()
            prix_element=driver.find_element_by_xpath("//*[@id='course_page']/div[1]/div[1]/section[2]/header/h1/strong")
            prix= u""+prix_element.text.strip()
            print("prix: ", prix)

            tr=driver.find_element_by_xpath("//div[@class='popover-container']/table/thead/tr")
            # print(tr.text)
            if str(tr.text).find("Jockey")==-1:
                # continue
                pass
            v1=l.split("/")[-1]
            v2=v1.split("-")
            texte=v2[-1]
            texte2="_".join(v2)
            texte3=texte2.replace("_"+texte,"")
            texte3=self.NettoyagePonctPresent(texte3).replace(" ","_").replace("\n","_").replace("\t","_")
            wb = Workbook()
            ws1 = wb.active

            # resultat = open(rep+"\\"+k+"_"+texte3+".txt", "w")
            #Ajout titre
            td_s=tr.find_elements_by_tag_name("th")
            liste_entete_web=[]
            liste=["Classement","Nom du cheval","Sexe Age","Poids","Chrono","PMU","Jockey","PALC","","Entraineur","CD","CF","Gains","Date","Hippodrome","Terrain","Prix","Ouvert a","Type","Allocation","Distance","","","Index"]
            for x in range(len(td_s)):
                titre=str(td_s[x].text).replace("\n"," ")
                liste_entete_web.append(titre)
            # enr="\t".join(liste)
            pos_entete_poids=0
            for b in range(len(liste_entete_web)):
                var2=u""+liste_entete_web[b].lower()
                if var2.find("poids")!=-1:
                    pos_entete_poids=b
                    break
            pos_entete_entraineur=0
            for b in range(len(liste_entete_web)):
                var2=u""+liste_entete_web[b].lower()
                if var2.find(u"entraîneur")!=-1:
                    pos_entete_entraineur=b
                    break
            pos_entete_cd=0
            for b in range(len(liste_entete_web)):
                var2=u""+liste_entete_web[b].lower()
                if var2.find(u"rapp. ouv.")!=-1:
                    pos_entete_cd=b
                    break
            pos_entete_cf=0
            for b in range(len(liste_entete_web)):
                var2=u""+liste_entete_web[b].lower()
                if var2.find(u"rapp.final pmu")!=-1:
                    pos_entete_cf=b
                    break
            pos_entete_jokey=0
            for b in range(len(liste_entete_web)):
                var2=u""+liste_entete_web[b].lower()
                if var2.find(u"jockey")!=-1:
                    pos_entete_jokey=b
                    break
            pos_entete_corde=0
            for b in range(len(liste_entete_web)):
                var2=u""+liste_entete_web[b].lower()
                if var2.find(u"corde")!=-1:
                    pos_entete_corde=b
                    break

            ws1.append(liste)
            # resultat.write(enr)
            #Ajout enregistrement
            tr_s=driver.find_elements_by_xpath("//div[@class='popover-container']/table/tbody/tr")
            ligne=1

            #Les valeurs fixes
            div=driver.find_element_by_xpath("//div[@class='row-fluid row-no-margin text-left']")
            p_s=div.find_elements_by_tag_name("p")
            t_prix=[]
            terrain=""
            type_var=""
            allocation=""
            distance=""
            metre=u"mètres"
            temps_total=""
            text_ouvert=""
            TypeDeCourse=""
            for w in range(len(p_s)):
                if p_s[w].text.find("Prix")!=-1:
                    p_text=p_s[w].text
                    a_partir_prix=p_text[int(p_text.find("Prix :")+len("Prix :")):]
                    pos1=a_partir_prix.find("\n")
                    pos2=a_partir_prix[1:].find("\n")
                    if pos2==-1:
                        prix_texte=a_partir_prix[pos1:].replace("\n","").replace("."," ")
                    else:
                        prix_texte=a_partir_prix[pos1:pos2+1].replace("\n","").replace("."," ")
                    t_prix=prix_texte.split(",")
                if p_s[w].text.find("Allocation")!=-1:
                    p_text=p_s[w].text
                    a_partir_prix=p_text[int(p_text.find("Allocation :")+len("Allocation :")):]
                    allocation_avec_s=False
                    if a_partir_prix.find("Allocation")!=-1:
                        a_partir_prix=p_text[int(p_text.find("Allocations :")+len("Allocations :")):]
                        allocation_avec_s=True
                    if allocation_avec_s==False:
                        pos1=a_partir_prix.find("\n")
                        pos2=a_partir_prix[1:].find("\n")
                        texte_allocation=a_partir_prix[pos1:pos2+1].replace("\n","").replace(".","")
                    else:
                        texte_allocation=a_partir_prix.strip("\n").strip()
                    #-----------allocation
                    compile_var="^.*([0-9.]*) (€|\$|£|yen|rouble)s? - .*$"
                    regex = re.compile(compile_var, re.IGNORECASE)
                    texte=texte_allocation
                    res =regex.findall(texte)
                    res_1=""
                    allocation=""
                    if len(res)>0:
                        res_1=res[0][0]
                        allocation = res_1.strip().replace(".", "")
                    else:
                        allocation=texte.strip().replace(".", "")
                    if texte.find("$")!=-1:
                        allocation = float(allocation.replace("$", "").strip())
                        allocation = allocation * 0.698
                    elif texte.find(" yen") !=-1:
                        allocation = float(allocation.replace(" yen", "").strip())
                        allocation = allocation * 0.0075
                    elif texte.find("rouble")!=-1:
                        allocation = float(allocation.replace("rouble", "").strip())
                        allocation = allocation * 0.023
                    elif texte.find("£")!=-1:
                        allocation = float(allocation.replace("£", "").strip())
                        allocation = allocation * 1.0196

                    allocation = str(allocation).replace("€", "").strip()
                    allocation=str(round(float(allocation), 2))
                    if str(allocation)[-2:]==".0":
                        allocation=str(allocation).replace(".0","")
                    print("allocation ", allocation.encode("cp1252"))


                if p_s[w].text.lower().find("terrain")!=-1:
                    t3= p_s[w].text.split("-")
                    type_var=t3[0]
                    for e in range(len(t3)):
                        if t3[e].lower().find("terrain")!=-1:
                            terrain_var=t3[e].split("\n")
                            if len(terrain_var)==0:
                                terrain=t3[e].replace("terrain","").replace("Terrain","").strip().capitalize()
                            else:
                                terrain=terrain_var[0].replace("terrain","").replace("Terrain","").strip().capitalize()

                if p_s[w].text.lower().find("temps total")!=-1:
                    t3= p_s[w].text.split("-")
                    for e in range(len(t3)):
                        if t3[e].lower().find("temps total")!=-1:
                            terrain_var=t3[e].split("\n")
                            if len(terrain_var)==0:
                                temps_total=t3[e].replace("temps total","").replace("Temps total","").strip()
                            else:
                                temps_total=terrain_var[0].replace("temps total","").replace("Temps total","").strip()

                if p_s[w].text.lower().find(metre)!=-1:
                    t3= p_s[w].text.split("-")
                    for e in range(len(t3)):
                        if t3[e].lower().find(metre)!=-1:
                            t8=t3[e].replace(".","").strip().split(" ")
                            distance=t8[0].strip()

                if p_s[w].text.find("\nPour ")!=-1 or p_s[w].text.find(". Pour ")!=-1:
                    p_text=p_s[w].text
                    pos1=p_text.find("\nPour ")
                    text_ouvert=p_text[pos1+1:]
                    #Ouvert
                    annee=datetime.now().year
                    #---------------conditions
                    conditions=""
                    compile_var="^.*\. Pour (.*)\..*poids(.*)$"
                    regex = re.compile(compile_var, re.IGNORECASE)
                    texte=text_ouvert
                    res =regex.findall(texte)
                    if len(res)>0:
                        conditions=res[0][0].strip()
                    else:
                        conditions=texte

                    M = False
                    F = False

                    if conditions.find("tous chevaux") !=-1:
                        M = True
                        F = True
                    if conditions.find("poulains")!=-1 or conditions.find("chevaux")!=-1 or conditions.find("chevaux entiers")!=-1 or conditions.find("hongres")!=-1:
                        M = True

                    if conditions.find("pouliches")!=-1 or conditions.find("juments")!=-1:
                        F = True

                    ConditionsEpure=""

                    if M and not F:
                        ConditionsEpure = "Males "
                    if F and not M:
                        ConditionsEpure = "Femelles "
                    if F and M:
                        ConditionsEpure = "M.H.F "

                    print("Conditions: " + conditions.encode("cp1252"))

                    NouvelleAnnee = True
                    if NouvelleAnnee == True:
                        compile_var="^(.*) née*s en ([0-9]*)( et antérieurement)*(.*)$"
                        regex = re.compile(compile_var, re.IGNORECASE)
                        texte=text_ouvert
                        res =regex.findall(texte)
                        res_1 = ""
                        res_2=""
                        res_3 = ""
                        res_4 = ""
                        if len(res)>0:
                            res_1=res[0][0]
                            res_2=res[0][1]
                            res_3=res[0][2]
                            res_4=res[0][3]
                            try:
                                Test = int(annee) - int(res_2)
                                if len(res_3) > 1:
                                    Test2 = " et au-dessus"
                                conditions = res_1+" de " + str(Test) + " ans" + Test2 + ", " + res_4
                            except:
                                pass

                    compile_var="(\d à \d{1,2}|\d\d|\d) ans"
                    regex = re.compile(compile_var, re.IGNORECASE)
                    texte=conditions
                    res =regex.findall(texte)
                    res_1=""

                    if len(res)>0:
                        res_1=res[0]
                        ConditionsEpure = ConditionsEpure + res_1 + " ans"
                    else:
                        ConditionsEpure = ConditionsEpure + conditions + " ans"
                    if conditions.find("au-dessus") !=-1:
                        ConditionsEpure = ConditionsEpure + " et +"
                    conditions = ConditionsEpure

                    TypeDeCourse = conditions.strip()
                    print("TypeDeCourse: ", TypeDeCourse.encode("cp1252"))


            for y in range(len(tr_s)):
                ligne=ligne+1
                td_s=tr_s[y].find_elements_by_tag_name("td")
                liste=[]
                col=0
                # chrono_1er=driver.find_element_by_xpath("//div[@class='row-fluid row-no-margin text-left']/p/strong[contains(text(), 'Temps total')]")
                # prix=driver.find_element_by_xpath("//div[@class='row-fluid row-no-margin text-left']/p/strong[contains(text(), 'Prix')]")

                for x in range(len(td_s)):
                    col=col+1
                    if pos_entete_poids>0 and x==pos_entete_poids:
                        d = ws1.cell(row=ligne, column=4, value=self.removeAccent(u""+td_s[x].text).strip()) #poids
                    elif pos_entete_corde>0 and x==pos_entete_corde: #corde
                        d = ws1.cell(row=ligne, column=8, value=self.removeAccent(u""+td_s[x].text).strip())
                    elif x==0:#classement
                         d = ws1.cell(row=ligne, column=1, value=self.removeAccent(u""+td_s[x].text).strip())
                    elif x==2:#pmu
                         d = ws1.cell(row=ligne, column=6, value=self.removeAccent(u""+td_s[x].text).strip())
                    elif x==3:#cheval
                         #Enlever les extensions
                         compile_var=u"^.*( .*)$".encode("utf8")
                         regex = re.compile(compile_var, re.IGNORECASE)
                         texte=u""+td_s[x].text.encode("utf8")
                         res =regex.findall(texte)
                         res_1=""
                         res_2=""
                         cheval_var=""
                         if len(res)>0:
                             res_1=res[0].strip().upper()
                             res_2=res[0]
                             if res_1 in code_pays:
                                 cheval_var=texte.replace(res_2,"").strip()
                         if cheval_var=="":
                            d = ws1.cell(row=ligne, column=2, value=self.removeAccent(u""+td_s[x].text).strip())
                         else:
                            d = ws1.cell(row=ligne, column=2, value=self.removeAccent(u""+cheval_var).strip())
                    elif x==4:#s/a
                         d = ws1.cell(row=ligne, column=3, value=self.removeAccent(u""+td_s[x].text).strip())
                    elif x==8:#chrono
                        if ligne==2 and u""+td_s[x].text=="":
                            d = ws1.cell(row=ligne, column=5, value=self.removeAccent(u""+temps_total).strip())
                        else:
                            d = ws1.cell(row=ligne, column=5, value=self.removeAccent(u""+td_s[x].text).strip())
                    # elif x==6:#poids
                    #     d = ws1.cell(row=ligne, column=4, value=u""+td_s[x].text)
                    elif pos_entete_jokey>0 and x==pos_entete_jokey:#jokey
                        d = ws1.cell(row=ligne, column=7, value=self.removeAccent(u""+td_s[x].text).strip())
                    elif pos_entete_entraineur>0 and x==pos_entete_entraineur:#entraineur
                        d = ws1.cell(row=ligne, column=10, value=self.removeAccent(u""+td_s[x].text).strip())
                    elif pos_entete_cd>0 and x==pos_entete_cd:#cd
                        d = ws1.cell(row=ligne, column=11, value=self.removeAccent(u""+td_s[x].text).strip())
                    elif pos_entete_cf>0 and x==pos_entete_cf:#cf
                        d = ws1.cell(row=ligne, column=12, value=self.removeAccent(u""+td_s[x].text).strip())
                    # else:
                    #     d = ws1.cell(row=ligne, column=col, value=str(td_s[x].text).encode("cp1252"))
                #gains
                try:
                    d = ws1.cell(row=ligne, column=13, value=self.removeAccent(u""+t_prix[y].replace(".","").replace(u"€","").strip()).strip())
                except:
                    pass
                #date
                d = ws1.cell(row=ligne, column=14, value=u""+date_jour)
                #hippodrome
                t1=k.split("_")
                if len(t1)==2:
                    hippodrome=t1[1].capitalize()
                elif len(t1)==3:
                    hippodrome=t1[1].capitalize()+" "+t1[2].capitalize()
                else:
                    hippodrome=""
                d = ws1.cell(row=ligne, column=15, value=self.removeAccent(u""+hippodrome).strip())
                #terrain
                d = ws1.cell(row=ligne, column=16, value=self.removeAccent(u""+terrain).strip())
                #prix
                d = ws1.cell(row=ligne, column=17, value=self.removeAccent(u""+prix).strip())
                #ouvert
                d = ws1.cell(row=ligne, column=18, value=self.removeAccent(u""+TypeDeCourse).strip())
                #type
                d = ws1.cell(row=ligne, column=19, value=self.removeAccent(u""+type_var).strip())
                #allocation
                d = ws1.cell(row=ligne, column=20, value=self.removeAccent(u""+allocation).strip())
                #distance
                d = ws1.cell(row=ligne, column=21, value=self.removeAccent(u""+distance).strip())


                # liste.append(str(td_s[x].text))

            if(os.access(rep+"\\"+k+"_"+texte3+".xlsx",os.F_OK)==False):
                wb.save(rep+"\\"+k+"_"+texte3+".xlsx")
            else:
                if distance=="":
                    wb.save(rep+"\\"+k+"_"+texte3+" "+str(num_fichier)+".xlsx")
                else:
                    wb.save(rep+"\\"+k+"_"+texte3+" "+str(distance)+".xlsx")


            try:
                driver.close()
            except:
                pass

            if os.path.exists('scraping_courses.lock')==True:
                os.remove('scraping_courses.lock')

            print("FIN traitement scrapping course")
            sys.exit(0)

        except Exception as inst:
            log=open(date_jour.replace("/", "-")+".txt", "a")
            traceback.print_exc(file=log)
            log.close()
            try:
                driver.close()
            except:
                pass
            sys.exit(0)

    def removeAccent(self,chaine,is2Maj=False):
        """ Cette fonction enleve les accents dans une chaine"""
        a=chaine
        a=a.replace("é","e")
        a=a.replace("è","e")
        a=a.replace("ê","e")
        a=a.replace("ë","e")

        a=a.replace("â","a")
        a=a.replace("à","a")
        a=a.replace("ä","a")

        a=a.replace("ü","u")
        a=a.replace("û","u")
        a=a.replace("ù","u")

        a=a.replace("ï","i")
        a=a.replace("î","i")
        a=a.replace("ö","o")
        a=a.replace("ô","o")
        a=a.replace("ç","c")

        a=a.replace("ñ","n")
        a=a.replace("œ","oe")

        a=a.replace("¨","")
        a=a.replace("^","")


        a=a.replace("Ë","E")
        a=a.replace("É","E")
        a=a.replace("Ê","E")
        a=a.replace("È","E")
        a=a.replace("Ä","A")
        a=a.replace("À","A")
        a=a.replace("Â","A")
        a=a.replace("Ü","U")
        a=a.replace("Ù","U")
        a=a.replace("Û","U")
        a=a.replace("Ï","I")
        a=a.replace("Î","I")
        a=a.replace("ÖÔÇ","O")
        a=a.replace("ÖÔÇ","O")
        a=a.replace("ÖÔÇ","C")
        a=a.replace("È","E")
        a=a.replace("&","&")
        if is2Maj==True:
            a=a.upper()
        return a

    def date2fr(self,sdateEn,sep="-"):
        a1=sdateEn[0:4]
        m1=sdateEn[5:7]
        d1=sdateEn[8:10]
        return d1+sep+m1+sep+a1

    def NettoyagePonctPresent(self,chaine):
        """ Cette fonction enleve les accents dans une chaine"""
        if chaine.find("<?>")==-1:
            ListeAccents = "ËÉÊÈÄÀÂÜÙÛÏÎÖÔÇëéêèäàâüùûïîöôç¨^,;:*!'-_.\\/?\"<>"
            ReplaceListeAccents = "EEEEAAAUUUIIOOCeeeeaaauuuiiooc                 "
            k=0
            chainenew=""
            while(k<len(chaine)):

                if (ord(chaine[k])>= 65 and ord(chaine[k]) <= 90) or (ord(chaine[k]) >= 48 and ord(chaine[k]) <= 57) or ord(chaine[k]) == 32:
                    pass
                else:
                    j=0
                    btrouve=False
                    while(j<len(ListeAccents)):
                        if(ListeAccents[j]==chaine[k]):
                            chaine  = chaine.replace(chaine[k],ReplaceListeAccents[j])
                            btrouve=True
                            break
                        j=j+1
                    if btrouve==False:
                        pass
                        #chaine  = chaine.replace(chaine[k]," ")

                k=k+1
            chaine = chaine.upper()
            return chaine.strip()
        else:
            return chaine


if __name__ == "__main__":
    Traitement()
